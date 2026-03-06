from flask import render_template, request, redirect, session, flash, url_for
from werkzeug.security import generate_password_hash
from werkzeug.utils import secure_filename
import os

from app.extensions import login_required, get_cursor
from app.org.blueprint import org_bp

from openpyxl import Workbook
from flask import send_file
import io

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from openpyxl.drawing.image import Image as ExcelImage
from reportlab.platypus import Image



from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import uuid


# =====================================================
# FILE UPLOAD CONFIG
# =====================================================
UPLOAD_FOLDER = "app/static/uploads/drivers"
ALLOWED_EXTENSIONS = {"pdf", "jpg", "jpeg", "png"}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# =====================================================
# DRIVER LIST
# =====================================================
@org_bp.route("/drivers")
@login_required
def org_drivers():
    db, cursor = get_cursor()

    cursor.execute("""
        SELECT 
            u.id,
            u.name,
            
            u.phone,

            d.driver_code,
            d.license_number,
            d.status,
            d.photo_path,
            d.license_type,
            d.experience_years,
            d.license_expiry,
            d.blood_group,
            d.emergency_contact,
            d.monthly_salary,

            b.bus_number,
            r.route_name,

            a.assignment,
            a.assignment_date,

            docs.license_pdf,
            docs.aadhar_pdf,
            docs.medical_pdf

        FROM users u
        JOIN driver_details d ON d.driver_id = u.id

        LEFT JOIN driver_assignment a
            ON a.id = (
                SELECT id FROM driver_assignment
                WHERE driver_id = u.id AND status='ASSIGNED'
                ORDER BY assignment_date DESC, id DESC
                LIMIT 1
            )

        LEFT JOIN buses b ON b.id = a.bus_id
        LEFT JOIN routes r ON r.id = a.route_id

        LEFT JOIN (
            SELECT
                driver_id,
                MAX(CASE WHEN document_type='LICENSE' THEN file_path END) AS license_pdf,
                MAX(CASE WHEN document_type='AADHAR' THEN file_path END) AS aadhar_pdf,
                MAX(CASE WHEN document_type='MEDICAL' THEN file_path END) AS medical_pdf
            FROM driver_documents
            GROUP BY driver_id
        ) docs ON docs.driver_id = u.id

        WHERE u.org_id=%s AND u.role='driver'
        ORDER BY u.name
    """, (session["org_id"],))

    drivers = cursor.fetchall()
    cursor.close()
    db.close()

    return render_template("org/org_driver_manage.html", drivers=drivers)


# =====================================================
# ADD DRIVER
# =====================================================
@org_bp.route("/drivers/add", methods=["GET", "POST"])
@login_required
def add_driver():
    if request.method == "POST":
        db, cursor = get_cursor()

        # -------- SAFETY --------
        monthly_salary = request.form.get("monthly_salary")
        if not monthly_salary:
            flash("Monthly salary is required", "error")
            return redirect("/org/drivers/add")

        # -------- CREATE USER --------
        cursor.execute("""
            INSERT INTO users (org_id, name, email, phone, role, password_hash)
            VALUES (%s,%s,%s,%s,'driver',%s)
        """, (
            session["org_id"],
            request.form["name"],
            request.form["email"],
            request.form["phone"],
            generate_password_hash(request.form["password"])
        ))

        driver_id = cursor.lastrowid
        
        # -------- DRIVER PHOTO UPLOAD --------
        photo = request.files.get("photo")

        photo_path = None

        if photo and photo.filename and allowed_file(photo.filename):
            filename = secure_filename(f"{driver_id}_photo_{photo.filename}")
            save_path = os.path.join(UPLOAD_FOLDER, filename)
            photo.save(save_path)

            photo_path = f"uploads/drivers/{filename}"



        # -------- DRIVER DETAILS --------
        cursor.execute("""
            INSERT INTO driver_details (
                driver_id, driver_code, license_number,
                license_type, experience_years,
                license_expiry, blood_group,
                emergency_contact, monthly_salary, status, photo_path
            )
             VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,'ACTIVE',%s)
        """, (
            driver_id,
            request.form["driver_code"],
            request.form["license_number"],
            request.form.get("license_type"),
            request.form.get("experience_years"),
            request.form["license_expiry"],
            request.form.get("blood_group"),
            request.form.get("emergency_contact"),
            monthly_salary,
            photo_path
        ))

        # -------- DOCUMENT UPLOAD --------
        documents = {
            "license_pdf": "LICENSE",
            "aadhar_pdf": "AADHAR",
            "medical_pdf": "MEDICAL"
        }

        for field, doc_type in documents.items():
            file = request.files.get(field)
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(f"{driver_id}_{doc_type}_{file.filename}")
                save_path = os.path.join(UPLOAD_FOLDER, filename)
                file.save(save_path)

                cursor.execute("""
                    INSERT INTO driver_documents
                        (org_id, driver_id, document_type, file_path)
                    VALUES (%s,%s,%s,%s)
                """, (
                    session["org_id"],
                    driver_id,
                    doc_type,
                    f"uploads/drivers/{filename}"
                ))

        db.commit()
        cursor.close()
        db.close()

        flash("Driver added successfully", "success")
        return redirect("/org/drivers")

    return render_template("org/add_driver.html")


# =====================================================
# EDIT DRIVER (DETAILS + DOCUMENTS)
# =====================================================
@org_bp.route("/drivers/<int:id>/edit", methods=["GET", "POST"])
@login_required
def edit_driver(id):
    db, cursor = get_cursor()

    cursor.execute("""
        SELECT
            u.name, u.email, u.phone,
            d.driver_code, d.license_number,
            d.license_type, d.experience_years,
            d.license_expiry, d.blood_group,
            d.emergency_contact, d.monthly_salary,
            d.photo_path
        FROM users u
        JOIN driver_details d ON d.driver_id=u.id
        WHERE u.id=%s AND u.org_id=%s
    """, (id, session["org_id"]))

    driver = cursor.fetchone()
    if not driver:
        cursor.close()
        db.close()
        return redirect("/org/drivers")

    if request.method == "POST":
        cursor.execute("""
            UPDATE users
            SET name=%s, email=%s, phone=%s
            WHERE id=%s AND org_id=%s
        """, (
            request.form["name"],
            request.form["email"],
            request.form["phone"],
            id,
            session["org_id"]
        ))

        cursor.execute("""
            UPDATE driver_details
            SET license_number=%s,
                license_type=%s,
                experience_years=%s,
                license_expiry=%s,
                blood_group=%s,
                emergency_contact=%s,
                monthly_salary=%s
            WHERE driver_id=%s
        """, (
            request.form["license_number"],
            request.form.get("license_type"),
            request.form.get("experience_years"),
            request.form["license_expiry"],
            request.form.get("blood_group"),
            request.form.get("emergency_contact"),
            request.form.get("monthly_salary", 0),
            id
        ))
        documents = {
            "license_pdf": "LICENSE",
            "aadhar_pdf": "AADHAR",
            "medical_pdf": "MEDICAL"
        }

        for field, doc_type in documents.items():
            file = request.files.get(field)

            if file and file.filename and allowed_file(file.filename):

                # 🔹 Get old document path
                cursor.execute("""
                    SELECT file_path FROM driver_documents
                    WHERE driver_id=%s AND document_type=%s
                """, (id, doc_type))

                old_doc = cursor.fetchone()

                if old_doc and old_doc.get("file_path"):
                    old_path = os.path.join("app/static", old_doc["file_path"])
                    if os.path.exists(old_path):
                        os.remove(old_path)

                # 🔹 Generate UNIQUE filename using UUID
                ext = file.filename.rsplit('.', 1)[1].lower()
                filename = f"{id}_{doc_type}_{uuid.uuid4().hex}.{ext}"

                save_path = os.path.join(UPLOAD_FOLDER, filename)
                file.save(save_path)

                new_path = f"uploads/drivers/{filename}"

                # 🔹 Update DB
                cursor.execute("""
                    INSERT INTO driver_documents
                        (org_id, driver_id, document_type, file_path)
                    VALUES (%s,%s,%s,%s)
                    ON DUPLICATE KEY UPDATE file_path=VALUES(file_path)
                """, (
                    session["org_id"],
                    id,
                    doc_type,
                    new_path
                ))
                
                
        # ----------------------------
        # UPDATE DRIVER PHOTO (IF NEW UPLOADED)
        # ----------------------------
        photo = request.files.get("photo")

        if photo and photo.filename and allowed_file(photo.filename):

            # 🔹 Get old photo
            cursor.execute("SELECT photo_path FROM driver_details WHERE driver_id=%s", (id,))
            old = cursor.fetchone()

            if old and old.get("photo_path"):
                old_file_path = os.path.join("app/static", old["photo_path"])
                if os.path.exists(old_file_path):
                    os.remove(old_file_path)

            # 🔹 Save new photo
            filename = secure_filename(f"{id}_photo_{photo.filename}")
            save_path = os.path.join(UPLOAD_FOLDER, filename)
            photo.save(save_path)

            new_photo_path = f"uploads/drivers/{filename}"

            # 🔹 Update DB
            cursor.execute("""
                UPDATE driver_details
                SET photo_path=%s
                WHERE driver_id=%s
            """, (new_photo_path, id))


        db.commit()
        cursor.close()
        db.close()

        flash("Driver updated successfully", "success")
        return redirect("/org/drivers")

    cursor.close()
    db.close()
    return render_template("org/edit_driver.html", driver=driver, driver_id=id)


# =====================================================
# RESET PASSWORD
# =====================================================
@org_bp.route("/drivers/<int:id>/reset-password", methods=["POST"])
@login_required
def reset_driver_password(id):
    db, cursor = get_cursor()

    cursor.execute("""
        UPDATE users
        SET password_hash=%s
        WHERE id=%s AND org_id=%s AND role='driver'
    """, (
        generate_password_hash(request.form["new_password"]),
        id,
        session["org_id"]
    ))

    db.commit()
    cursor.close()
    db.close()

    flash("Password reset successfully", "success")
    return redirect("/org/drivers")


# =====================================================
# TOGGLE DRIVER STATUS
# =====================================================
@org_bp.route("/drivers/<int:id>/toggle")
@login_required
def toggle_driver(id):
    db, cursor = get_cursor()

    cursor.execute("""
        UPDATE driver_details
        SET status =
            CASE
                WHEN status='ACTIVE' THEN 'INACTIVE'
                WHEN status='INACTIVE' THEN 'SUSPENDED'
                ELSE 'ACTIVE'
            END
        WHERE driver_id=%s
    """, (id,))

    db.commit()
    cursor.close()
    db.close()

    return redirect("/org/drivers")


# =====================================================
# DELETE DRIVER
# =====================================================
@org_bp.route("/drivers/<int:id>/delete", methods=["POST"])
@login_required
def delete_driver(id):
    db, cursor = get_cursor()

    cursor.execute("DELETE FROM driver_documents WHERE driver_id=%s", (id,))
    cursor.execute("DELETE FROM driver_details WHERE driver_id=%s", (id,))
    cursor.execute(
        "DELETE FROM users WHERE id=%s AND org_id=%s",
        (id, session["org_id"])
    )

    db.commit()
    cursor.close()
    db.close()

    flash("Driver deleted", "success")
    return redirect("/org/drivers")

# --------------------------
# DOWNLOAD DRIVER LIST (EXCEL)
# --------------------------
@org_bp.route("/drivers/download/excel")
@login_required
def download_drivers_excel():
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, numbers
    from flask import send_file
    import io

    db, cursor = get_cursor()

    cursor.execute("""
        SELECT 
            u.name,
            u.phone,
            d.driver_code,
            d.license_number,
            d.status,
            d.license_type,
            d.experience_years,
            d.license_expiry,
            d.blood_group,
            d.emergency_contact,
            d.monthly_salary,
            d.photo_path
        FROM users u
        JOIN driver_details d ON d.driver_id = u.id
        WHERE u.org_id=%s AND u.role='driver'
        ORDER BY u.name
    """, (session["org_id"],))

    drivers = cursor.fetchall()
    cursor.close()
    db.close()

    # ---------- CREATE EXCEL ----------
    wb = Workbook()
    ws = wb.active
    ws.title = "Drivers"

    headers = [
        "Photo","Name", "Phone", "Driver Code", "License No", "Status",
        "License Type", "Experience (Years)", "License Expiry",
        "Blood Group", "Emergency Contact", "Monthly Salary"
    ]

    ws.append(headers)

    # ---------- STYLE HEADER ----------
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # ---------- ADD DATA ----------
    row_number = 2

    for d in drivers:

        # Add empty photo cell
        ws.cell(row=row_number, column=1, value="")

        ws.cell(row=row_number, column=2, value=d.get("name", ""))
        ws.cell(row=row_number, column=3, value=d.get("phone", ""))
        ws.cell(row=row_number, column=4, value=d.get("driver_code", ""))
        ws.cell(row=row_number, column=5, value=d.get("license_number", ""))
        ws.cell(row=row_number, column=6, value=d.get("status", ""))
        ws.cell(row=row_number, column=7, value=d.get("license_type", ""))
        ws.cell(row=row_number, column=8, value=d.get("experience_years", ""))
        ws.cell(row=row_number, column=9, value=d.get("license_expiry", ""))
        ws.cell(row=row_number, column=10, value=d.get("blood_group", ""))
        ws.cell(row=row_number, column=11, value=d.get("emergency_contact", ""))
        ws.cell(row=row_number, column=12, value=d.get("monthly_salary", ""))

        # Add photo if exists
        photo_path = d.get("photo_path")
        if photo_path:
            full_path = os.path.join("app/static", photo_path)
            if os.path.exists(full_path):
                img = ExcelImage(full_path)
                img.width = 60
                img.height = 60
                ws.add_image(img, f"A{row_number}")

        ws.row_dimensions[row_number].height = 45
        row_number += 1

    # ---------- FORMAT DATE COLUMN (License Expiry = Column 8) ----------
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            if cell.value:
                cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

    # ---------- FORMAT SALARY COLUMN (Column 11) ----------
    for row in ws.iter_rows(min_row=2, min_col=11, max_col=11):
        for cell in row:
            if cell.value:
                cell.number_format = '#,##0'

    # ---------- AUTO COLUMN WIDTH ----------
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 3

    # ---------- SEND FILE ----------
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name="drivers_list.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
# ----------------------------
# DOWNLOAD DRIVER LIST (PDF)
# ----------------------------
@org_bp.route("/drivers/download/pdf")
@login_required
def download_drivers_pdf():
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer
    )
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from flask import send_file
    from datetime import datetime
    import io

    db, cursor = get_cursor()

    cursor.execute("""
        SELECT
            d.photo_path,
            d.driver_code,
            u.name,
            u.phone,
            d.license_number,
            d.license_type,
            d.experience_years,
            d.status
        FROM users u
        JOIN driver_details d ON d.driver_id = u.id
        WHERE u.org_id=%s AND u.role='driver'
        ORDER BY u.name
    """, (session["org_id"],))

    drivers = cursor.fetchall()
    cursor.close()
    db.close()

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=36,
        rightMargin=36,
        topMargin=40,
        bottomMargin=36
    )

    styles = getSampleStyleSheet()
    elements = []

    # ---------- HEADER ----------
    title_style = ParagraphStyle(
        "Title",
        fontSize=20,
        alignment=1,
        textColor=colors.HexColor("#1e3a8a"),
        spaceAfter=6
    )

    subtitle_style = ParagraphStyle(
        "Subtitle",
        fontSize=10,
        alignment=1,
        textColor=colors.grey,
        spaceAfter=12
    )

    elements.append(Paragraph("🚍 Driver List Report", title_style))
    elements.append(Paragraph(
        f"Generated on: {datetime.now().strftime('%d %b %Y, %I:%M %p')}",
        subtitle_style
    ))

    # Divider line
    elements.append(Table(
        [[""]],
        colWidths=[doc.width],
        style=[("LINEBELOW", (0, 0), (-1, -1), 1, colors.grey)]
    ))

    elements.append(Spacer(1, 18))

    # ---------- TABLE DATA ----------
    data = [[
        "Photo",
        "Driver Code",
        "Name",
        "Phone",
        "License No",
        "License Type",
        "Experience",
        "Status"
    ]]

    for d in drivers:

        photo_cell = "N/A"

        photo_path = d.get("photo_path")
        if photo_path:
            full_path = os.path.join("app/static", photo_path)
            if os.path.exists(full_path):
                photo_cell = Image(full_path, width=40, height=40)

        data.append([
            photo_cell,
            d.get("driver_code", ""),
            d.get("name", ""),
            d.get("phone", ""),
            d.get("license_number", ""),
            d.get("license_type", ""),
            f"{d.get('experience_years', '')} yrs",
            d.get("status", "")
        ])

    # ---------- TABLE (FULL WIDTH) ----------
    table = Table(
        data,
        repeatRows=1,
        colWidths=[
            doc.width * 0.10,  # Photo
            doc.width * 0.12,  # Driver Code
            doc.width * 0.18,  # Name
            doc.width * 0.14,  # Phone
            doc.width * 0.16,  # License No
            doc.width * 0.14,  # License Type
            doc.width * 0.12,  # Experience
            doc.width * 0.10   # Status
        ]
    )

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),

        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),

        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
        ("TOPPADDING", (0, 1), (-1, -1), 6),
    ]))

    elements.append(table)

    # ---------- FOOTER ----------
    elements.append(Spacer(1, 22))
    elements.append(Paragraph(
        "This document is system generated and does not require a signature.",
        ParagraphStyle(
            "Footer",
            fontSize=9,
            alignment=1,
            textColor=colors.grey
        )
    ))

    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="drivers_list.pdf",
        mimetype="application/pdf"
    )
